#   Copyright 2022 NEC Corporation
#
#   Licensed under the Apache License, Version 2.0 (the "License");
#   you may not use this file except in compliance with the License.
#   You may obtain a copy of the License at
#
#       http://www.apache.org/licenses/LICENSE-2.0
#
#   Unless required by applicable law or agreed to in writing, software
#   distributed under the License is distributed on an "AS IS" BASIS,
#   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#   See the License for the specific language governing permissions and
#   limitations under the License.
import requests_mock

import io
import openpyxl
from zipfile import ZipFile, BadZipfile

import os
import ulid
import json
import time
import datetime
import re
from contextlib import closing

from unittest import mock
import threading
from importlib import import_module
import copy

from common_library.common import const
from common_library.common import encrypt
from common_library.common.db import DBconnector
from common_library.common import bl_plan_service, api_keycloak_users

import job_manager
import job_manager_const
import job_manager_config
from jobs import jobs_common
from jobs.UserImportJobExecutor import UserImportJobExecutor

from tests.common import test_common

from libs.job_manager_classes import SubProcessesManager
from common_library.common import user_import_file_common

USERNAME_COL_INDEX = 2
ERROR_TEXT_COL_INDEX = 12

def test_execute_registration_nomally():
    """ユーザー登録正常系 / User registration normal pattern
    """
    testdata = import_module("tests.db.exports.testdata")

    with test_common.requsts_mocker_default():

        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        queue = make_queue_import_user('ja', organization_id, user_import_data=[data_sample_registration])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、登録成功件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 1 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 0
        assert t["SUCCESS_REGISTER"] == 1 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 0 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0


def test_execute_delete_nomally():
    """ユーザー削除正常系 / User delete normal pattern
    """
    testdata = import_module("tests.db.exports.testdata")

    with test_common.requsts_mocker_default():

        organization_id = list(testdata.ORGANIZATIONS.keys())[0]

        # 登録
        queue = make_queue_import_user('ja', organization_id, user_import_data=[data_sample_registration])
        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 削除
        queue = make_queue_import_user('ja', organization_id, user_import_data=[data_sample_delete], job_type=const.JOB_TYPE_USER_BULK_DELETE)
        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、削除成功件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 0 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 1
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 1
        assert t["FAILED_REGISTER"] == 0 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0


def test_execute_file_error():
    """ファイル異常パターン
    """
    testdata = import_module("tests.db.exports.testdata")

    # ファイルの形式エラーパターン
    with test_common.requsts_mocker_default():

        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        queue = make_queue_import_user('ja', organization_id, user_import_data_image=b'1234567890')

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 失敗を返すこと
        assert not result
        # 状態が異常であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_FAILED
        assert t["MESSAGE"] == 'Excelファイルの読み込みに失敗しました。'

    # 列数超過エラーパターン
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        lang = 'ja'

        # 書き換えたExcelを生成
        imp_wb = user_import_file_common.UserResultWorkbook(lang, error_column=False)
        imp_wb.ws.cell(1, job_manager_config.JOBS[const.PROCESS_KIND_USER_IMPORT]["extra_config"]["max_number_of_cols_allowd"]+1).value = "dummy"
        excel_image = imp_wb.get_workbook_bytes_image()

        queue = make_queue_import_user(lang, organization_id, user_import_data_image=excel_image)
        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 失敗を返すこと
        assert not result
        # 状態が異常であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_FAILED
        assert t["MESSAGE"].startswith('Excelファイルの列数が多すぎます。')

    # 行数超過エラーパターン
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        lang = 'ja'

        # 書き換えたExcelを生成
        imp_wb = user_import_file_common.UserResultWorkbook(lang, error_column=False)
        imp_wb.ws.cell(job_manager_config.JOBS[const.PROCESS_KIND_USER_IMPORT]["extra_config"]["max_number_of_rows_allowd"]+1,1).value = "dummy"
        excel_image = imp_wb.get_workbook_bytes_image()

        queue = make_queue_import_user(lang, organization_id, user_import_data_image=excel_image)
        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 失敗を返すこと
        assert not result
        # 状態が異常であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_FAILED
        assert t["MESSAGE"].startswith('Excelファイルの行数が多すぎます。')

    # ヘッダーセルの形式（文言）エラーパターン
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        lang = 'ja'

        # ヘッダを書き換えたExcelを生成
        imp_wb = user_import_file_common.UserResultWorkbook(lang, error_column=False)
        imp_wb.ws.cell(1,1).value = 'dummy-header'
        excel_image = imp_wb.get_workbook_bytes_image()

        queue = make_queue_import_user(lang, organization_id, user_import_data_image=excel_image)
        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 失敗を返すこと
        assert not result
        # 状態が異常であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_FAILED
        assert t["MESSAGE"] == 'Excelファイルに必須の項目がありません。({0})'.format('実行処理種別')

def test_execute_validation_error():
    """Validation error pattern
    """
    testdata = import_module("tests.db.exports.testdata")

    # ユーザー名エラー
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        err_username = "_err_user"
        queue = make_queue_import_user('ja', organization_id, user_import_data=[{**data_sample_registration, **{"USERNAME": err_username}}])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        # 成功を返すこと
        assert result

        # 状態が完了で、登録失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 1 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 0
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 1 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0

        # 結果シートの中身が正しいこと
        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, USERNAME_COL_INDEX).value == err_username
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value.startswith("先頭の文字にアルファベット以外が指定されています。")

        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

    # パスワードエラー
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        queue = make_queue_import_user('ja', organization_id, user_import_data=[{**data_sample_registration, **{"PASSWORD": ""}}])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、登録失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 1 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 0
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 1 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, USERNAME_COL_INDEX).value == data_sample_registration["USERNAME"]
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value.startswith("必須項目が不足しています。")

    # emailエラー
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        queue = make_queue_import_user('ja', organization_id, user_import_data=[{**data_sample_registration, **{"EMAIL": "error-format"}}])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、登録失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 1 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 0
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 1 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, USERNAME_COL_INDEX).value == data_sample_registration["USERNAME"]
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value.startswith("メールアドレスの形式に誤りがあります。")

    # firstnameエラー
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        queue = make_queue_import_user('ja', organization_id, user_import_data=[{**data_sample_registration, **{"FIRSTNAME": "a".ljust(const.length_user_firstName + 1, "_")}}])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、登録失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 1 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 0
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 1 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, USERNAME_COL_INDEX).value == data_sample_registration["USERNAME"]
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value.startswith("指定可能な文字数を超えています。(項目:{0}".format("名"))

    # lastnameエラー
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        queue = make_queue_import_user('ja', organization_id, user_import_data=[{**data_sample_registration, **{"LASTNAME": "a".ljust(const.length_user_lastName + 1, "_")}}])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、登録失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 1 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 0
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 1 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, USERNAME_COL_INDEX).value == data_sample_registration["USERNAME"]
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value.startswith("指定可能な文字数を超えています。(項目:{0}".format("姓"))

    # 有効エラー
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        queue = make_queue_import_user('ja', organization_id, user_import_data=[{**data_sample_registration, **{"ENABLED": "ERROR-VALUE"}}])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、登録失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 1 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 0
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 1 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, USERNAME_COL_INDEX).value == data_sample_registration["USERNAME"]
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value.startswith("True/False 以外が指定されています。")

    # 所属エラー
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        queue = make_queue_import_user('ja', organization_id, user_import_data=[{**data_sample_registration, **{"AFFILIATION": "a".ljust(const.length_user_affiliation + 1, "_")}}])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、登録失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 1 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 0
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 1 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, USERNAME_COL_INDEX).value == data_sample_registration["USERNAME"]
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value.startswith("指定可能な文字数を超えています。(項目:{0}".format("所属"))

    # 説明エラー
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        queue = make_queue_import_user('ja', organization_id, user_import_data=[{**data_sample_registration, **{"DESCRIPTION": "a".ljust(const.length_user_description + 1, "_")}}])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、登録失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 1 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 0
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 1 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, USERNAME_COL_INDEX).value == data_sample_registration["USERNAME"]
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value.startswith("指定可能な文字数を超えています。(項目:{0}".format("説明"))

    # ロールエラー
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        error_role = "ERROR-ROLE"
        queue = make_queue_import_user('ja', organization_id, user_import_data=[{**data_sample_registration, **{"ROLES": data_sample_registration["ROLES"] + "," + error_role}}])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、登録失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 1 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 0
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 1 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, USERNAME_COL_INDEX).value == data_sample_registration["USERNAME"]
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value.startswith("指定されたロール({0})は存在しません".format(error_role))


def test_execute_registration_error_limits():
    """userのlimit超過エラー
    """
    testdata = import_module("tests.db.exports.testdata")

    # 上限数エラー
    limit_users = 2
    with test_common.requsts_mocker_default(), \
        mock.patch.object(bl_plan_service, "organization_limits_get", return_value={const.RESOURCE_COUNT_USERS: limit_users}): # limit値を設定

        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        queue = make_queue_import_user('ja', organization_id,
                    user_import_data=[
                        {**data_sample_registration, **{"USERNAME": "limit-test01", "EMAIL": "limit-test01@example.com"}},
                        {**data_sample_registration, **{"USERNAME": "limit-test02", "EMAIL": "limit-test02@example.com"}},
                        {**data_sample_registration, **{"USERNAME": "limit-test03", "EMAIL": "limit-test03@example.com"}},
                    ])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、登録失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 3 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 0
        assert t["SUCCESS_REGISTER"] == 1 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 2 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 2
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, USERNAME_COL_INDEX).value == "limit-test02"
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value == "{0}の上限数({1})を超えるため、新しい{0}は作成できません。".format("ユーザー", limit_users)
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 2, USERNAME_COL_INDEX).value == "limit-test03"
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 2, ERROR_TEXT_COL_INDEX).value == "{0}の上限数({1})を超えるため、新しい{0}は作成できません。".format("ユーザー", limit_users)

def test_execute_registration_error_duplicate():
    """重複ユーザーの登録
    """
    testdata = import_module("tests.db.exports.testdata")

    # 重複エラー（ユーザー名）
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        queue = make_queue_import_user('ja', organization_id,
                    user_import_data=[
                        {**data_sample_registration, **{"USERNAME": "duplicate-name-test01", "EMAIL": "duplicate-name-test01@example.com"}},
                        {**data_sample_registration, **{"USERNAME": "duplicate-name-test01", "EMAIL": "duplicate-name-test02@example.com"}},
                    ])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 失敗を返すこと
        assert not result

        # 状態が異常であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_FAILED
        assert t["MESSAGE"] == '対象のユーザー名が重複しています(重複ユーザー名:[{0}])'.format('duplicate-name-test01')

    # 重複エラー（email）
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        queue = make_queue_import_user('ja', organization_id,
                    user_import_data=[
                        {**data_sample_registration, **{"USERNAME": "duplicate-email-test01", "EMAIL": "duplicate-email-test01@example.com"}},
                        {**data_sample_registration, **{"USERNAME": "duplicate-email-test02", "EMAIL": "duplicate-email-test01@example.com"}},
                    ])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、登録成功・登録失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 2 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 0
        assert t["SUCCESS_REGISTER"] == 1 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 1 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, USERNAME_COL_INDEX).value == "duplicate-email-test02"
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value.startswith("指定されたユーザーはすでに存在しているため作成できません。")

def test_execute_registration_error_user_create():
    """ユーザー作成失敗
    """
    testdata = import_module("tests.db.exports.testdata")

    # keycloak HTTP-400応答
    with test_common.requsts_mocker_default() as requests_mocker:
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        keycloak_error_msg = "http-400 error"
        requests_mocker.register_uri(
            requests_mock.POST,
            re.compile(rf'^{test_common.keycloak_origin()}/auth/admin/realms/{organization_id}/users'),
            status_code=400,
            json={"errorMessage": keycloak_error_msg})

        queue = make_queue_import_user('ja', organization_id, user_import_data=[data_sample_registration])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、登録失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 1 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 0
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 1 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value == "ユーザー作成に失敗しました({0})".format(keycloak_error_msg)

    # keycloak HTTP-500応答
    with test_common.requsts_mocker_default() as requests_mocker:
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        keycloak_error_msg = "http-500 error"
        requests_mocker.register_uri(
            requests_mock.POST,
            re.compile(rf'^{test_common.keycloak_origin()}/auth/admin/realms/{organization_id}/users'),
            status_code=500,
            json={"errorMessage": keycloak_error_msg})

        queue = make_queue_import_user('ja', organization_id, user_import_data=[data_sample_registration])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、登録失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 1 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 0
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 1 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value == "ユーザー作成に失敗しました(対象ユーザー:{0})".format(data_sample_registration["USERNAME"])

    # keycloak user作成後、user取得失敗(http-500応答)
    with test_common.requsts_mocker_default() as requests_mocker:
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        error_user = "get_failed_pattern1"
        keycloak_error_msg = "http-500 error"
        requests_mocker.register_uri(
            requests_mock.GET,
            re.compile(rf'^{test_common.keycloak_origin()}/auth/admin/realms/{organization_id}/users\?username={error_user}'),
            status_code=500,
            json={"errorMessage": keycloak_error_msg})

        queue = make_queue_import_user('ja', organization_id, user_import_data=[{**data_sample_registration, **{"USERNAME": error_user}}])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、登録失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 1 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 0
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 1 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value.startswith("ユーザー作成に失敗しました")


    # keycloak user作成後、user取得失敗(Userなし)
    with test_common.requsts_mocker_default() as requests_mocker:
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        requests_mocker.register_uri(
            requests_mock.POST,
            re.compile(rf'^{test_common.keycloak_origin()}/auth/admin/realms/{organization_id}/users'),
            status_code=201,
            json={})

        queue = make_queue_import_user('ja', organization_id, user_import_data=[{**data_sample_registration, **{"USERNAME": "get_failed_pattern2"}}])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、登録失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 1 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 0
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 1 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value.startswith("ユーザー作成に失敗しました")

    # 登録と削除が混在の場合
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]

        # ユーザー削除処理
        queue = make_queue_import_user('ja', organization_id, user_import_data=[data_sample_registration2, data_sample_delete])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 失敗を返すこと
        assert not result

        # 状態が異常であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_FAILED
        assert t["MESSAGE"] == 'ユーザー一括インポートの場合、実行処理種別に「登録」「更新」以外は指定できません。'

def test_execute_delete_error_user_delete():
    """ユーザー削除失敗
    """
    testdata = import_module("tests.db.exports.testdata")

    # keycloak HTTP-500応答(ユーザー情報取得に失敗)
    with test_common.requsts_mocker_default() as requests_mocker:
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        keycloak_error_msg = "http-500 error"
        requests_mocker.register_uri(
            requests_mock.GET,
            re.compile(rf'^{test_common.keycloak_origin()}/auth/admin/realms/{organization_id}/users'),
            status_code=500,
            json={"errorMessage": keycloak_error_msg})

        queue = make_queue_import_user('ja', organization_id, user_import_data=[data_sample_delete], job_type=const.JOB_TYPE_USER_BULK_DELETE)

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、削除失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 0 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 1
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 0 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 1

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value == "ユーザーの取得に失敗しました(対象ユーザー:{0})".format(data_sample_registration["USERNAME"])

    # 存在しないユーザーを削除
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]

        # ユーザー削除処理
        queue = make_queue_import_user('ja', organization_id, user_import_data=[data_sample_delete], job_type=const.JOB_TYPE_USER_BULK_DELETE)

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、削除失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 0 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 1
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 0 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 1

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value == "指定されたユーザーが存在しません"

    # keycloak HTTP-500応答(ロール情報情報取得に失敗)
    with test_common.requsts_mocker_default() as requests_mocker:
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        # ユーザー作成登録
        user_id = add_user(user=user_json1)

        # 削除対象のユーザーを取得 / Get target user
        u_get = get_user(user_json1["username"])
        u_get_json = json.loads(u_get.text)
        user_id = u_get_json[0]["id"]

        # ユーザー削除処理
        organization_private = DBconnector().get_organization_private(organization_id)
        keycloak_error_msg = "http-500 error"
        client_id = organization_private.user_token_client_id
        requests_mocker.register_uri(
            requests_mock.GET,
            re.compile(rf'^{test_common.keycloak_origin()}/auth/admin/realms/{organization_id}/users/{user_id}/role-mappings/clients/{client_id}/composite'),
            status_code=500,
            json={"errorMessage": keycloak_error_msg})

        queue = make_queue_import_user('ja', organization_id, user_import_data=[data_sample_delete], job_type=const.JOB_TYPE_USER_BULK_DELETE)

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、削除失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 0 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 1
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 0 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 1

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value == "ロールの取得に失敗しました(対象ID:{0} client:{1})".format(organization_id, client_id)

    # オーガナイゼーション管理者ロールのユーザーを削除
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]

        # ユーザー削除処理
        queue = make_queue_import_user('ja', organization_id, user_import_data=[data_sample_delete_admin], job_type=const.JOB_TYPE_USER_BULK_DELETE)

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、削除失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 0 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 1
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 0 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 1

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value == "オーガナイゼーション管理者ロールのユーザーは削除できません"

    # keycloak HTTP-500応答(ユーザー削除に失敗)
    with test_common.requsts_mocker_default() as requests_mocker:
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        # ユーザー作成登録
        user_id = add_user(user=user_json1)

        # 削除対象のユーザーを取得 / Get target user
        u_get = get_user(user_json1["username"])
        u_get_json = json.loads(u_get.text)
        user_id = u_get_json[0]["id"]

        # ユーザー削除処理
        organization_private = DBconnector().get_organization_private(organization_id)
        keycloak_error_msg = "http-500 error"
        requests_mocker.register_uri(
            requests_mock.DELETE,
            re.compile(rf'^{test_common.keycloak_origin()}/auth/admin/realms/{organization_id}/users/{user_id}'),
            status_code=500,
            json={"errorMessage": keycloak_error_msg})

        queue = make_queue_import_user('ja', organization_id, user_import_data=[data_sample_delete], job_type=const.JOB_TYPE_USER_BULK_DELETE)

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、削除失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 0 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 1
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 0 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 1

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value == "ユーザー削除に失敗しました(対象ユーザーID:{0})".format(user_id)

    # 削除と登録が混在の場合
    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]

        # ユーザー削除処理
        queue = make_queue_import_user('ja', organization_id, user_import_data=[data_sample_registration, data_sample_delete2], job_type=const.JOB_TYPE_USER_BULK_DELETE)

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 失敗を返すこと
        assert not result

        # 状態が異常であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_FAILED
        assert t["MESSAGE"] == 'ユーザー一括削除の場合、実行処理種別に「削除」以外は指定できません。'

def test_execute_registration_error_role_mapping():
    """ロール付与失敗
    """
    testdata = import_module("tests.db.exports.testdata")

    # keycloak role mapping失敗
    with test_common.requsts_mocker_default() as requests_mocker:
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        requests_mocker.register_uri(
            requests_mock.POST,
            re.compile(rf'^{test_common.keycloak_origin()}/auth/admin/realms/{organization_id}/users/[^/]*/role-mappings/clients/'),
            status_code=500,
            json={})

        queue = make_queue_import_user('ja', organization_id, user_import_data=[{**data_sample_registration, **{"USERNAME": "failed_role_mapping_pattern", "EMAIL": "failed_role_mapping_pattern@example.com"}}])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 成功を返すこと
        assert result

        # 状態が完了で、登録失敗件数が１であること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_COMP
        assert t["COUNT_REGISTER"] == 1 and t["COUNT_UPDATE"] == 0 and t["COUNT_DELETE"] == 0
        assert t["SUCCESS_REGISTER"] == 0 and t["SUCCESS_UPDATE"] == 0 and t["SUCCESS_DELETE"] == 0
        assert t["FAILED_REGISTER"] == 1 and t["FAILED_UPDATE"] == 0 and t["FAILED_DELETE"] == 0

        ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
        assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
        assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value.startswith("ロール設定に失敗しました")

def test_execute_error_not_found_jobs_table():
    """TABLEの情報なし
    """
    testdata = import_module("tests.db.exports.testdata")

    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        queue = make_queue_import_user('ja', organization_id, user_import_data=[data_sample_registration])

        with closing(DBconnector().connect_orgdb(organization_id)) as conn,\
            conn.cursor() as cursor:
            cursor.execute("DELETE FROM T_JOBS_USER WHERE JOB_ID = %(job_id)s", {"job_id": queue["PROCESS_EXEC_ID"]})
            conn.commit()

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 失敗を返すこと
        assert not result

    with test_common.requsts_mocker_default():
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        queue = make_queue_import_user('ja', organization_id, user_import_data=[data_sample_registration])

        with closing(DBconnector().connect_orgdb(organization_id)) as conn,\
            conn.cursor() as cursor:
            cursor.execute("DELETE FROM T_JOBS_USER_FILE WHERE JOB_ID = %(job_id)s", {"job_id": queue["PROCESS_EXEC_ID"]})
            conn.commit()

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 失敗を返すこと
        assert not result
        # 状態が失敗で、メッセージが設定されていること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_FAILED
        assert t["MESSAGE"] == "処理対象のレコードの取得に失敗しました(テーブル:{0})".format("T_JOBS_USER_FILE")

def test_execute_error_get_specifiable_roles():
    """オーガナイゼーションのロール一覧取得失敗
    """
    testdata = import_module("tests.db.exports.testdata")

    with test_common.requsts_mocker_default() as requests_mocker:
        organization_id = list(testdata.ORGANIZATIONS.keys())[0]
        requests_mocker.register_uri(
            requests_mock.GET,
            re.compile(rf'^{test_common.keycloak_origin()}/auth/admin/realms/{organization_id}/clients/[^/]*/roles'),
            status_code=500,
            json={})

        queue = make_queue_import_user('ja', organization_id, user_import_data=[{**data_sample_registration, **{"USERNAME": "failed_role_mapping_pattern", "EMAIL": "failed_role_mapping_pattern@example.com"}}])

        executor = UserImportJobExecutor(queue)
        result = executor.execute_base()
        save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        # 失敗を返すこと
        assert not result
        # 状態が失敗で、メッセージが設定されていること
        t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
        assert t["JOB_STATUS"] == const.JOB_USER_FAILED
        assert t["MESSAGE"].startswith("ロールの取得に失敗しました")


def test_execute_timeout():
    """Timeoutによる中断
    """
    testdata = import_module("tests.db.exports.testdata")

    organization_id = list(testdata.ORGANIZATIONS.keys())[0]

    timeout_sec = 5

    # Jobの設定を試験用に切り替え
    process_kind=const.PROCESS_KIND_USER_IMPORT
    job_config_jobs = copy.deepcopy(job_manager_config.JOBS)
    job_config_jobs[process_kind]["timeout_seconds"] = timeout_sec
    job_config_jobs[process_kind]["extra_config"]["status_update_interval"] = 3

    with test_common.requsts_mocker_default(), mock.patch.dict(f"job_manager_config.JOBS", job_config_jobs):

        # sub process起動用の情報生成
        sub_processes_mgr = SubProcessesManager()
        sub_process_parameter = sub_processes_mgr.generate_sub_process_parameter()
        sub_processes_mgr.append_new_process(sub_process_parameter, test_common.get_main_process(os.getpid()))

        # sub processの起動
        main_thread = threading.Thread(
            target=job_manager.job_manager_sub_process,
            args=(sub_process_parameter,),
            name="sub_process",
            daemon=True
        )
        main_thread.start()
        try:
            user_import_data = [{**data_sample_registration, **{"USERNAME": f'test{i}', "EMAIL": f"test{i}@example.com"}} for i in range(1,1001)]
            queue = make_queue_import_user('ja', organization_id, user_import_data=user_import_data)

            # timeout発生まで待つ
            time.sleep(timeout_sec)
            for i in range(30):
                t = select_t_jobs_user(organization_id, queue["PROCESS_EXEC_ID"])
                if t["JOB_STATUS"] != const.JOB_USER_EXEC:
                    break
                time.sleep(1)

            # タイムアウトエラーで終了していること
            assert t["JOB_STATUS"] == const.JOB_USER_FAILED
            assert '行目の処理中にタイムアウトしました。' in t["MESSAGE"]

            # 結果ファイルができていること
            ws = get_result_worksheet(organization_id, queue["PROCESS_EXEC_ID"])
            assert ws.max_row == user_import_file_common.EXCEL_HEADER_ROWS + 1
            assert ws.cell(user_import_file_common.EXCEL_HEADER_ROWS + 1, ERROR_TEXT_COL_INDEX).value == 'タイムアウト発生のため、この行を処理中に中断しました。'

            save_result_file(organization_id, queue["PROCESS_EXEC_ID"])

        finally:
            # sub processの終了要求
            sub_processes_mgr.set_sub_process_termination_request(force=True)

            # sub processの終了確認
            # assert test_common.check_state(
            #     timeout=5.0, conditions=lambda : not main_thread.is_alive())

            main_thread.join()


def test_force_update_status_normally():
    """強制ステータス更新の確認
    """
    testdata = import_module("tests.db.exports.testdata")

    datas = [
        {
            "organization_id": list(testdata.ORGANIZATIONS.keys())[0],
            "queue_exists": False,
            "too_old_date": True,
            "queue": None,
        },
        {
            "organization_id": list(testdata.ORGANIZATIONS.keys())[0],
            "queue_exists": False,
            "too_old_date": True,
            "queue": None,
        },
        {
            "organization_id": list(testdata.ORGANIZATIONS.keys())[1],
            "queue_exists": False,
            "too_old_date": True,
            "queue": None,
        },
        {
            "organization_id": list(testdata.ORGANIZATIONS.keys())[1],
            "queue_exists": True,
            "too_old_date": True,
            "queue": None,
        },
        {
            "organization_id": list(testdata.ORGANIZATIONS.keys())[2],
            "queue_exists": False,
            "too_old_date": True,
            "queue": None,
        },
        {
            "organization_id": list(testdata.ORGANIZATIONS.keys())[2],
            "queue_exists": False,
            "too_old_date": False,
            "queue": None,
        },
    ]
    for data in datas:
        queue = make_queue_import_user('ja', data['organization_id'], user_import_data_image=b'dummy', insert_queue=data["queue_exists"])
        data['queue'] = queue

        with closing(DBconnector().connect_orgdb(data['organization_id'])) as conn,\
            conn.cursor() as cursor:

            if data["too_old_date"]:
                cursor.execute("""
                    UPDATE T_JOBS_USER SET LAST_UPDATE_TIMESTAMP = DATE_SUB(LAST_UPDATE_TIMESTAMP, INTERVAL %(SECONDS)s SECOND)
                        WHERE JOB_ID = %(JOB_ID)s
                """, {"JOB_ID": queue["PROCESS_EXEC_ID"], "SECONDS": job_manager_config.JOBS[job_manager_const.PROCESS_KIND_FORCE_UPDATE_STATUS]['extra_config']['prograss_seconds'] + 1})
                conn.commit()

    # 強制ステータス更新を実行
    UserImportJobExecutor.force_update_status()
    # データの更新状況を確認
    for i, data in enumerate(datas):
        t = select_t_jobs_user(data['organization_id'], data["queue"]["PROCESS_EXEC_ID"])

        if data['queue_exists']:
            # queueに情報が残っている場合は、日時とは関係なく更新しない
            assert t["JOB_STATUS"] == const.JOB_USER_NOT_EXEC
        elif not data['too_old_date']:
            # 日時が古くない場合は、更新しない
            assert t["JOB_STATUS"] == const.JOB_USER_NOT_EXEC
        else:
            # queueにデータがなく、日付が古い場合、エラーに更新していること
            assert t["JOB_STATUS"] == const.JOB_USER_FAILED


def make_queue_import_user(lang, organization_id, user_import_data=None, user_import_data_image=None, insert_queue=True, job_type=const.JOB_TYPE_USER_BULK_IMPORT):
    """テスト用共通:Queue情報生成

    Args:
        lang (_type_): _description_
        organization_id (_type_): _description_
        user_import_data (_type_, optional): _description_. Defaults to None.
        user_import_data_image (_type_, optional): _description_. Defaults to None.
        insert_queue (bool, optional): _description_. Defaults to True.

    Returns:
        _type_: _description_
    """
    process_id = ulid.new().str
    process_exec_id = ulid.new().str
    file_id = ulid.new().str

    queue={
        "PROCESS_ID": process_id,
        "PROCESS_KIND": const.PROCESS_KIND_USER_IMPORT,
        "PROCESS_EXEC_ID": process_exec_id,
        "ORGANIZATION_ID": organization_id,
        "WORKSPACE_ID": None,
        "LAST_UPDATE_USER": job_manager_const.SYSTEM_USER_ID,
        "LAST_UPDATE_TIMESTAMP": str(datetime.datetime.now()),
    }
    if user_import_data is not None:
        excel_image = make_excel_bytes_image(user_import_data, lang)
    else:
        excel_image = user_import_data_image

    with closing(DBconnector().connect_orgdb(organization_id)) as conn,\
        conn.cursor() as cursor:

        cursor.execute('''
            INSERT INTO T_JOBS_USER
            (JOB_ID,
            JOB_TYPE,
            JOB_STATUS,
            LANGUAGE)
            VALUES
            (%(JOB_ID)s,
            %(JOB_TYPE)s,
            %(JOB_STATUS)s,
            %(LANGUAGE)s)
            ''',
            {"JOB_ID": process_exec_id,
            "JOB_TYPE": job_type,
            "JOB_STATUS": const.JOB_USER_NOT_EXEC,
            "LANGUAGE": lang}
        )

        cursor.execute('''
            INSERT INTO T_JOBS_USER_FILE
            (FILE_ID,
            JOB_ID,
            FILE_DATA)
            VALUES
            (%(FILE_ID)s,
            %(JOB_ID)s,
            %(FILE_DATA)s)
            ''',
            {"FILE_ID": file_id,
            "JOB_ID": process_exec_id,
            "FILE_DATA": excel_image}
            )

        if os.environ.get("USER_EXCEL_FILE_SAVED") == "TRUE":
            with open(f'/workspace/platform_root/platform_job/tests/temp/{process_exec_id}.xlsx', 'wb') as fp:
                fp.write(excel_image)


        conn.commit()

    if insert_queue:
        test_common.insert_queue([queue])

    return queue

def get_result_worksheet(organization_id, job_id):
    """テスト用共通:結果情報Excel worksheet取得

    Args:
        organization_id (_type_): _description_
        job_id (_type_): _description_

    Returns:
        _type_: _description_
    """
    with closing(DBconnector().connect_orgdb(organization_id)) as conn,\
        conn.cursor() as cursor:

        cursor.execute("SELECT * FROM T_JOBS_USER_RESULT WHERE JOB_ID = %(job_id)s", {"job_id": job_id})
        row = cursor.fetchone()
        excel_zipfile = ZipFile(io.BytesIO(row["FILE_DATA"])).fp
        wb = openpyxl.load_workbook(excel_zipfile, read_only=True, data_only=True, keep_links=False, rich_text=False)
        return wb.worksheets[0]

def save_result_file(organization_id, job_id):
    """テスト用共通:結果情報Excelファイル保存

    Args:
        organization_id (_type_): _description_
        job_id (_type_): _description_
    """
    if os.environ.get("USER_EXCEL_FILE_SAVED") == "TRUE":
        with closing(DBconnector().connect_orgdb(organization_id)) as conn,\
            conn.cursor() as cursor:

            cursor.execute("SELECT * FROM T_JOBS_USER_RESULT WHERE JOB_ID = %(job_id)s", {"job_id": job_id})
            row = cursor.fetchone()
            if row is not None:
                with open(f'{os.environ.get("TEST_OUTPUT_PATH")}/{job_id}_result.xlsx', 'wb') as fp:
                    fp.write(row["FILE_DATA"])


def make_excel_bytes_image(user_import_data, lang):
    """テスト用共通:Import Excel ファイルイメージ生成

    Args:
        user_import_data (_type_): _description_
        lang (_type_): _description_

    Returns:
        _type_: _description_
    """
    imp_wb = user_import_file_common.UserResultWorkbook(lang, error_column=False)

    for row in user_import_data:
        imp_wb.write_row(row)

    return imp_wb.get_workbook_bytes_image()


def select_t_jobs_user(organization_id, job_id):
    """テスト用共通:T_JOBS_USER取得

    Args:
        organization_id (_type_): _description_
        job_id (_type_): _description_

    Returns:
        _type_: _description_
    """
    with closing(DBconnector().connect_orgdb(organization_id)) as conn,\
        conn.cursor() as cursor:

        cursor.execute("SELECT * FROM T_JOBS_USER WHERE JOB_ID = %(job_id)s", {"job_id": job_id})
        return cursor.fetchone()


def add_user(user):
        """ユーザの追加 / add user

        Args:
            cell_values (dict): cellの値 / cell value

        Returns:
            str: user id
        """
        testdata = import_module("tests.db.exports.testdata")

        with closing(DBconnector().connect_orgdb(list(testdata.ORGANIZATIONS.keys())[0])) as conn:
            organization_private = DBconnector().get_organization_private(list(testdata.ORGANIZATIONS.keys())[0])
            organization_sa_token = jobs_common.organization_sa_token(list(testdata.ORGANIZATIONS.keys())[0], organization_private)

        # ユーザーの追加 / add user
        u_create = api_keycloak_users.user_create(
            realm_name=list(testdata.ORGANIZATIONS.keys())[0], user_json=user, token=organization_sa_token.get()
        )
        return u_create

def get_user(user_name):
        """ユーザ情報取得 / get user

        Args:
            user_name (str): ユーザー名 / user name

        Returns:
            dict: u_get
        """
        testdata = import_module("tests.db.exports.testdata")

        with closing(DBconnector().connect_orgdb(list(testdata.ORGANIZATIONS.keys())[0])) as conn:
            organization_private = DBconnector().get_organization_private(list(testdata.ORGANIZATIONS.keys())[0])
            organization_sa_token = jobs_common.organization_sa_token(list(testdata.ORGANIZATIONS.keys())[0], organization_private)

        # ユーザーの取得 / get user
        u_get = api_keycloak_users.user_get(realm_name=list(testdata.ORGANIZATIONS.keys())[0], user_name=user_name, token=organization_sa_token.get())

        return u_get


# サンプルデータ（追加用）
data_sample_registration = {
    "PROC_TYPE": "追加",
    "USERNAME": "testuser-01",
    "PASSWORD": "password",
    "EMAIL": "testuser-01@example.com",
    "LASTNAME": "testuser",
    "FIRSTNAME": "01",
    "ENABLED": "TRUE",
    "AFFILIATION": "所属01",
    "DESCRIPTION": "説明01",
    "USER_ID": None,
    "ROLES": "_ws1-admin",
}

data_sample_registration2 = {
    "PROC_TYPE": "追加",
    "USERNAME": "testuser-02",
    "PASSWORD": "password",
    "EMAIL": "testuser-02@example.com",
    "LASTNAME": "testuser",
    "FIRSTNAME": "02",
    "ENABLED": "TRUE",
    "AFFILIATION": "所属01",
    "DESCRIPTION": "説明01",
    "USER_ID": None,
    "ROLES": "_ws1-admin",
}

# サンプルデータ（削除用）
data_sample_delete = {
    "PROC_TYPE": "削除",
    "USERNAME": "testuser-01",
}
data_sample_delete2 = {
    "PROC_TYPE": "削除",
    "USERNAME": "testuser-02",
}

data_sample_delete_admin = {
    "PROC_TYPE": "削除",
    "USERNAME": "admin",
}


# サンプルユーザーデータ
user_json1 = {
    "username": "testuser-01",
    "email": "testuser-01@example.com",
    "firstName": "01",
    "lastName": "testuser",
    "credentials": [
        {
            "type": "password",
            "value": "password",
            "temporary": True,
        }
    ],
    "attributes":
    {
        "affiliation": "所属01",
        "description": "説明01",
    },
    "enabled": True
}