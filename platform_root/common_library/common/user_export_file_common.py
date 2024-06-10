#   Copyright 2024 NEC Corporation
#
#   Licensed under the Apache License, Version 2.0 (the "License");
#   you may not use this file except in compliance with the License.
#   You may obtain a copy of the License at
#
#       http://www.apache.org/licenses/LICENSE-2.0
#
#   Unless required by applicable law or agreed to in writing, software
#   distributed under the License is distributed on an "AS IS" BASIS,
#   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#   See the License for the specific language governing permissions and
#   limitations under the License.
import io
import openpyxl
import time

from common_library.common import const
from common_library.common import multi_lang

import job_manager_config

# Excelのヘッダーの行数 / Number of rows in Excel header
EXCEL_HEADER_ROWS = 2

# COLUMN情報 / COLUMN information
COLUMN_IDS = {
    "USERNAME": {
        "text": "ユーザー名",
        "text-id": "000-62006",
        "description": None,
        "description-id": "000-62007",
        "width": 18,
    },
    "PASSWORD": {
        "text": "パスワード",
        "text-id": "000-62008",
        "description": None,
        "description-id": "000-62009",
        "width": 16,
    },
    "EMAIL": {
        "text": "email",
        "text-id": "000-62010",
        "description": None,
        "description-id": "000-62011",
        "width": 30,
    },
    "LASTNAME": {
        "text": "姓",
        "text-id": "000-62012",
        "description": None,
        "description-id": "000-62013",
        "width": 12,
    },
    "FIRSTNAME": {
        "text": "名",
        "text-id": "000-62014",
        "description": None,
        "description-id": "000-62015",
        "width": 12,
    },
    "ENABLED": {
        "text": "有効",
        "text-id": "000-62016",
        "description": "TRUE/FALSE",
        "description-id": "000-62017",
        "width": 9,
    },
    "AFFILIATION": {
        "text": "所属",
        "text-id": "000-62018",
        "description": None,
        "description-id": "000-62019",
        "width": 16,
    },
    "DESCRIPTION": {
        "text": "説明",
        "text-id": "000-62020",
        "description": None,
        "description-id": "000-62021",
        "width": 16,
    },
    "ROLES": {
        "text": "ロール",
        "text-id": "000-62024",
        "description": "カンマ区切りで列挙",
        "description-id": "000-62025",
        "width": 28,
    },
    "USER_ID": {
        "text": "ユーザーID",
        "text-id": "000-62022",
        "description": "追加時は不要",
        "description-id": "000-62023",
        "width": 38,
    },
}

class UserResultWorkbook():
    """結果出力用Excelファイル生成Class / Excel file generation class for result output
    """
    def __init__(self, lang, error_column=False):
        """constructor

        Args:
            lang (str): Language
            error_column (bool, optional): エラー項目生成有無 / Whether error items are generated or not. Defaults to False.
        """
        self.lang = lang
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.worksheets[0]
        self.col_indexes = {}

        # 項目をid, 項目位置でdictionary化する / Convert items to dictionary by id and item position
        if error_column:
            column_ids = COLUMN_IDS
        else:
            column_ids = {cid: item for cid, item in COLUMN_IDS.items() if cid != "ERROR_TEXT"}

        # ヘッダー行のスタイル生成 / Header row style generation
        header_row1_fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='00459D')
        header_row2_fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='D9D9D9')
        header_side = openpyxl.styles.borders.Side(style='thin', color='FFFFFF')
        header_border = openpyxl.styles.borders.Border(top=header_side, bottom=header_side, left=header_side, right=header_side)
        header_row1_alignment = openpyxl.styles.Alignment(wrapText=True, vertical = "center")
        header_row2_alignment = openpyxl.styles.Alignment(wrapText=True, vertical = "top")
        header_row1_font = openpyxl.styles.fonts.Font(color='FFFFFF')
        header_row2_font = openpyxl.styles.fonts.Font(color='000000')

        # ヘッダ行の生成 / Generate header row
        for idx, key in enumerate(column_ids):
            self.ws.column_dimensions[chr(ord('A')+idx)].width = column_ids[key]["width"]
            self.ws.cell(row = 1, column=idx+1).value = multi_lang.get_text_spec(lang, column_ids[key]["text-id"], column_ids[key]["text"])
            self.ws.cell(row = 1, column=idx+1).fill = header_row1_fill
            self.ws.cell(row = 1, column=idx+1).border = header_border
            self.ws.cell(row = 1, column=idx+1).alignment = header_row1_alignment
            self.ws.cell(row = 1, column=idx+1).font = header_row1_font
            self.ws.cell(row = 2, column=idx+1).value = multi_lang.get_text_spec(lang, column_ids[key]["description-id"], column_ids[key]["description"])
            self.ws.cell(row = 2, column=idx+1).fill = header_row2_fill
            self.ws.cell(row = 2, column=idx+1).border = header_border
            self.ws.cell(row = 2, column=idx+1).alignment = header_row2_alignment
            self.ws.cell(row = 2, column=idx+1).font = header_row2_font
            self.col_indexes[key] = idx

        # ヘッダ行の高さ設定 / Header row height setting
        self.ws.row_dimensions[1].height = 18
        self.ws.row_dimensions[2].height = 42
        # スクロール固定の設定 / Fixed scroll setting
        self.ws.freeze_panes = f'{chr(ord("A")+self.col_indexes["USERNAME"]+1)}{EXCEL_HEADER_ROWS+1}'

        # 書き込み行の初期設定 / Initial setting of write line
        self.__writed_row_idx = EXCEL_HEADER_ROWS

    def write_row(self, cell_values):
        """行の書き込み / writing a line

        Args:
            cell_values (dict): cellの値 / cell value
        """
        self.__writed_row_idx += 1
        for cid, cvalue in cell_values.items():
            self.ws.cell(self.__writed_row_idx, self.col_indexes[cid] + 1).value = cvalue

    def get_workbook_bytes_image(self):
        """Excelファイルイメージ取得 / Excel file image acquisition

        Returns:
            byte[]: Excel file image
        """
        return make_workbook_bytes_image(self.wb)

def make_workbook_bytes_image(wb):
    """Excelファイルイメージを生成する / Generate Excel file image

    Args:
        wb (openpyxl.Workbook): Excel Workbook

    Returns:
        byte[]: Excel file image
    """
    mem_stream = io.BytesIO()
    wb.save(mem_stream)
    mem_stream.seek(0)
    mem_buf = mem_stream.read()

    del mem_stream
    return mem_buf
