#   Copyright 2024 NEC Corporation
#
#   Licensed under the Apache License, Version 2.0 (the "License");
#   you may not use this file except in compliance with the License.
#   You may obtain a copy of the License at
#
#       http://www.apache.org/licenses/LICENSE-2.0
#
#   Unless required by applicable law or agreed to in writing, software
#   distributed under the License is distributed on an "AS IS" BASIS,
#   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#   See the License for the specific language governing permissions and
#   limitations under the License.
import io
import openpyxl
import collections
from zipfile import ZipFile

from common_library.common import multi_lang
from common_library.common.common import FileFormatErrorException


# Excelのヘッダーの行数 / Number of rows in Excel header
EXCEL_HEADER_ROWS = 2
# 明細の作成行 / Detail creation line
EXCEL_FORMAT_SET_ROWS = 100

# 実行処理種別 / Execution processing type
PROC_TYPE_ADD = ["追加", "Registration"]
PROC_TYPE_UPD = ["更新", "Update"]
PROC_TYPE_DEL = ["削除", "Delete"]
PROC_TYPES = PROC_TYPE_ADD + PROC_TYPE_UPD + PROC_TYPE_DEL

# COLUMN情報 / COLUMN information
COLUMN_IDS = {
    "PROC_TYPE": {
        "text": "実行処理種別",
        "text-id": "000-62001",
        "description": "追加/更新/削除",
        "description-id": "000-62002",
        "width": 14,
    },
    "USERNAME": {
        "text": "ユーザー名",
        "text-id": "000-62006",
        "description": None,
        "description-id": "000-62007",
        "width": 18,
    },
    "PASSWORD": {
        "text": "初期パスワード",
        "text-id": "000-62008",
        "description": "追加のみ有効",
        "description-id": "000-62009",
        "width": 16,
    },
    "EMAIL": {
        "text": "email",
        "text-id": "000-62010",
        "description": "他のユーザーと重複できません",
        "description-id": "000-62011",
        "width": 30,
    },
    "LASTNAME": {
        "text": "姓",
        "text-id": "000-62012",
        "description": None,
        "description-id": "000-62013",
        "width": 12,
    },
    "FIRSTNAME": {
        "text": "名",
        "text-id": "000-62014",
        "description": None,
        "description-id": "000-62015",
        "width": 12,
    },
    "ENABLED": {
        "text": "有効",
        "text-id": "000-62016",
        "description": "TRUE/FALSE",
        "description-id": "000-62017",
        "width": 9,
    },
    "AFFILIATION": {
        "text": "所属",
        "text-id": "000-62018",
        "description": None,
        "description-id": "000-62019",
        "width": 16,
    },
    "DESCRIPTION": {
        "text": "説明",
        "text-id": "000-62020",
        "description": None,
        "description-id": "000-62021",
        "width": 16,
    },
    "ROLES": {
        "text": "ロール",
        "text-id": "000-62024",
        "description": "カンマ区切りで列挙",
        "description-id": "000-62025",
        "width": 28,
    },
    "USER_ID": {
        "text": "ユーザーID",
        "text-id": "000-62022",
        "description": "参照のみ項目。\n一括処理では使用されません。",
        "description-id": "000-62023",
        "width": 38,
    },
    "ERROR_TEXT": {
        "text": "エラー内容",
        "text-id": "000-62026",
        "description": None,
        "description-id": "000-62027",
        "width": 80,
    },
}


class UserResultWorkbook():
    """結果出力用Excelファイル生成Class / Excel file generation class for result output
    """
    def __init__(self, lang, error_column=False):
        """constructor

        Args:
            lang (str): Language
            error_column (bool, optional): エラー項目生成有無 / Whether error items are generated or not. Defaults to False.
        """
        self.lang = lang
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.worksheets[0]
        self.col_indexes = {}

        # 項目をid, 項目位置でdictionary化する / Convert items to dictionary by id and item position
        if error_column:
            column_ids = COLUMN_IDS
        else:
            column_ids = {cid: item for cid, item in COLUMN_IDS.items() if cid != "ERROR_TEXT"}

        # ヘッダー行のスタイル生成 / Header row style generation
        header_row1_fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='00459D')
        header_row2_fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='D9D9D9')
        header_side = openpyxl.styles.borders.Side(style='thin', color='FFFFFF')
        header_border = openpyxl.styles.borders.Border(top=header_side, bottom=header_side, left=header_side, right=header_side)
        header_row1_alignment = openpyxl.styles.Alignment(wrapText=True, vertical="center")
        header_row2_alignment = openpyxl.styles.Alignment(wrapText=True, vertical="top")
        header_row1_font = openpyxl.styles.fonts.Font(color='FFFFFF')
        header_row2_font = openpyxl.styles.fonts.Font(color='000000')

        # ヘッダ行の生成 / Generate header row
        for idx, key in enumerate(column_ids):
            self.ws.column_dimensions[chr(ord('A') + idx)].width = column_ids[key]["width"]
            self.ws.cell(row=1, column=idx + 1).value = multi_lang.get_text_spec(lang, column_ids[key]["text-id"], column_ids[key]["text"])
            self.ws.cell(row=1, column=idx + 1).fill = header_row1_fill
            self.ws.cell(row=1, column=idx + 1).border = header_border
            self.ws.cell(row=1, column=idx + 1).alignment = header_row1_alignment
            self.ws.cell(row=1, column=idx + 1).font = header_row1_font
            self.ws.cell(row=2, column=idx + 1).value = multi_lang.get_text_spec(lang, column_ids[key]["description-id"], column_ids[key]["description"])
            self.ws.cell(row=2, column=idx + 1).fill = header_row2_fill
            self.ws.cell(row=2, column=idx + 1).border = header_border
            self.ws.cell(row=2, column=idx + 1).alignment = header_row2_alignment
            self.ws.cell(row=2, column=idx + 1).font = header_row2_font
            self.col_indexes[key] = idx

        # ヘッダ行の高さ設定 / Header row height setting
        self.ws.row_dimensions[1].height = 18
        self.ws.row_dimensions[2].height = 42
        # スクロール固定の設定 / Fixed scroll setting
        self.ws.freeze_panes = f'{chr(ord("A")+self.col_indexes["USERNAME"]+1)}{EXCEL_HEADER_ROWS+1}'

        # 書き込み行の初期設定 / Initial setting of write line
        self.__writed_row_idx = EXCEL_HEADER_ROWS

    def write_row(self, cell_values):
        """行の書き込み / writing a line

        Args:
            cell_values (dict): cellの値 / cell value
        """
        self.__writed_row_idx += 1
        for cid, cvalue in cell_values.items():
            self.ws.cell(self.__writed_row_idx, self.col_indexes[cid] + 1).value = cvalue

    def get_workbook_bytes_image(self):
        """Excelファイルイメージ取得 / Excel file image acquisition

        Returns:
            byte[]: Excel file image
        """
        # 入力エリアのcell設定はファイル出力時に設定 / Input area cell settings are set when outputting a file

        # cellの入力選択肢生成 / Cell input option generation
        proc_type_lang_idx = 0 if self.lang == "ja" else 1
        proc_type_col_validation = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            allow_blank=True,
            errorStyle="warning",
            formula1=f'"{PROC_TYPE_ADD[proc_type_lang_idx]},{PROC_TYPE_UPD[proc_type_lang_idx]},{PROC_TYPE_DEL[proc_type_lang_idx]}"'
        )
        enabled_col_validation = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            allow_blank=True,
            errorStyle="warning",
            formula1='"TRUE,FALSE"'
        )

        # PROC_TYPEのcellの入力選択肢設定 / PROC_TYPE cell input option settings
        proc_type_col = chr(ord('A') + self.col_indexes["PROC_TYPE"])
        proc_type_col_validation.add(f"{proc_type_col}{EXCEL_HEADER_ROWS+1}:{proc_type_col}{self.ws.max_row + EXCEL_FORMAT_SET_ROWS}")
        self.ws.add_data_validation(proc_type_col_validation)
        # ENABLEDのcellの入力選択肢設定 / ENABLED cell input option settings
        enabled_col = chr(ord('A') + self.col_indexes["ENABLED"])
        enabled_col_validation.add(f"{enabled_col}{EXCEL_HEADER_ROWS+1}:{enabled_col}{self.ws.max_row + EXCEL_FORMAT_SET_ROWS}")
        self.ws.add_data_validation(enabled_col_validation)

        return make_workbook_bytes_image(self.wb)


class UserImportWorkbook():
    """User Import用読み込みclass / Reading class for User Import
    """
    def __init__(self, lang, file_image, max_number_of_cols_allowd, max_number_of_rows_allowd, xl_buffered_rows):
        """constructor

        Args:
            lang (str): Language
            file_image (byte[]): Excel file image
            max_number_of_cols_allowd (int): Maximum number of columns in an Excel file that can be read
            max_number_of_rows_allowd (int): Maximum number of rows in an Excel file that can be read
            xl_buffered_rows (int): Excel file buffering row count
        """
        self.lang = lang
        self.max_number_of_cols_allowd = max_number_of_cols_allowd
        self.max_number_of_rows_allowd = max_number_of_rows_allowd
        self.xl_buffered_rows = xl_buffered_rows

        # Excelファイルイメージ取り込み / Excel file image import
        try:
            excel_zipfile = ZipFile(io.BytesIO(file_image)).fp
            self.wb = openpyxl.load_workbook(excel_zipfile, read_only=True, data_only=True, keep_links=False, rich_text=False)
        except Exception as ex:
            raise FileFormatErrorException(multi_lang.get_text_spec(self.lang, '401-00015', 'Excelファイルの読み込みに失敗しました。'))

        # 先頭シートを対象とする / Target the first sheet
        try:
            self.ws = self.wb.worksheets[0]
            if self.ws is None:
                raise Exception('no sheet')
        except Exception as ex:
            raise FileFormatErrorException(multi_lang.get_text_spec(self.lang, '401-00017', 'Excelファイルにワークシートが存在しません。'))

        # Excelの行列の最大数のチェック / Checking maximum number of columns in Excel
        if self.ws.max_column > self.max_number_of_cols_allowd:
            raise FileFormatErrorException(
                multi_lang.get_text_spec(
                    self.lang, '401-00018', 'Excelファイルの列数が多すぎます。（列数:{0},列数最大:{1})',
                    self.ws.max_column,
                    self.max_number_of_cols_allowd))

        if self.ws.max_row > self.max_number_of_rows_allowd:
            raise FileFormatErrorException(
                multi_lang.get_text_spec(
                    self.lang, '401-00019', 'Excelファイルの行数が多すぎます。（行数:{0},行数最大:{1})',
                    self.ws.max_column,
                    self.max_number_of_rows_allowd))

        # 先頭行を取得する / get first row
        ws_header = [self.ws.cell(row=1, column=col_idx).value for col_idx in range(1, self.ws.max_column + 1)]

        # ヘッダーのテキストよりidと位置をdictionary化する / Convert id and position into dictionary from header text
        self.col_indexes = {cid: search_list_value(ws_header, multi_lang.get_text_spec(lang, citem["text-id"], citem["text"]))
                            for cid, citem in COLUMN_IDS.items()}

        # ヘッダーの必須項目チェック / Check required items in header
        not_found_columns = [multi_lang.get_text_spec(lang, COLUMN_IDS[cid]["text-id"], COLUMN_IDS[cid]["text"])
                             for cid, index in self.col_indexes.items() if index == -1 and cid != "ERROR_TEXT"]
        if len(not_found_columns) > 0:
            raise FileFormatErrorException(multi_lang.get_text_spec(self.lang, '401-00016', 'Excelファイルに必須の項目がありません。({0})', ",".join(not_found_columns)))

        # Read用の項目: iter_rowsでxl_buffered_rowsずつ読み取りバッファリングして読み込む（性能対策）
        # Item for Read: Iter_rows reads xl_buffered_rows by buffering and reading (performance measure)
        self.__buff = []
        self.__buffered_row_idx = EXCEL_HEADER_ROWS
        self.__row_idx = EXCEL_HEADER_ROWS

    def count_proc_type(self):
        """処理件数のカウント / Counting the number of processed items

        Returns:
            int: 登録件数 / Number of registrations
            int: 変更件数 / Number of changes
            int: 削除件数 / Number of items deleted
        """
        count_register = count_update = count_delete = 0

        # 一つ一つ読みだすと遅いのでiter_rowsで一気に読み取る / Reading them one by one is slow, so read them all at once with iter_rows
        for row in self.ws.iter_rows(EXCEL_HEADER_ROWS + 1, self.ws.max_row, self.col_indexes["PROC_TYPE"], self.col_indexes["PROC_TYPE"]):
            for cel in row:
                if cel.value in PROC_TYPE_ADD:
                    count_register += 1
                elif cel.value in PROC_TYPE_UPD:
                    count_update += 1
                elif cel.value in PROC_TYPE_DEL:
                    count_delete += 1

        return count_register, count_update, count_delete

    def check_duplicate_user_name(self):
        """ユーザー名の重複チェック / Check for duplicate usernames

        Returns:
            list: 重複しているユーザー名 / list of duplicate users
        """
        user_list = []
        # 一つ一つ読みだすと遅いのでiter_rowsで一気に読み取る / Reading them one by one is slow, so read them all at once with iter_rows
        for row in self.ws.iter_rows(EXCEL_HEADER_ROWS + 1, self.ws.max_row, self.col_indexes["USERNAME"], self.col_indexes["USERNAME"]):
            for cel in row:
                # 空欄行の場合スキップ
                if not cel.value:
                    continue
                user_list.append(cel.value)

        # 重複しているユーザー名を抽出 / Extract duplicate usernames
        duplicate_user_list = [k for k, v in collections.Counter(user_list).items() if v > 1]

        return duplicate_user_list

    def read_row(self):
        """行の読み取り / reading row

        Returns:
            dict: cellの値 / cell value
        """
        if len(self.__buff) == 0:
            # bufferの件数が0件の時は、bufferへの読み込みを行う

            if self.__buffered_row_idx == self.ws.max_row:
                # 最終行までバッファ済みの場合は、もう読み取る行が無いのでNoneを返す
                # If the last line has been buffered, there are no more lines to read, so return None.
                return None

            # バッファリングする最終行を決定する / Determine the last line to buffer
            max_row = self.__buffered_row_idx + self.xl_buffered_rows
            if max_row > self.ws.max_row:
                max_row = self.ws.max_row

            # バッファリングの対象行をiter_rowsで指定しバッファにため込む
            # Specify the rows to be buffered with iter_rows and store them in the buffer
            for xl_row in self.ws.iter_rows(self.__buffered_row_idx + 1, max_row, 1, max(self.col_indexes.values())):
                cell_values = [xl_cel.value for xl_cel in xl_row]
                self.__buff.append(
                    {
                        cid: cell_values[col_idx - 1]
                        for cid, col_idx in self.col_indexes.items()
                        if cid != "ERROR_TEXT"
                    }
                )

            # バッファリングが済んだ行を設定 / Set buffered line
            self.__buffered_row_idx = max_row

        # バッファの先頭行を返し、先頭行はバッファから削除する / Returns the first line of the buffer and deletes the first line from the buffer
        row = self.__buff[0].copy()
        del self.__buff[0]

        self.__row_idx += 1
        return row

    def get_row_idx(self):
        """現在の処理行を返す / Returns the current processed row

        Returns:
            int: 現在の処理行 / current processed row
        """
        return self.__row_idx


def make_workbook_bytes_image(wb):
    """Excelファイルイメージを生成する / Generate Excel file image

    Args:
        wb (openpyxl.Workbook): Excel Workbook

    Returns:
        byte[]: Excel file image
    """
    mem_stream = io.BytesIO()
    wb.save(mem_stream)
    mem_stream.seek(0)
    mem_buf = mem_stream.read()

    del mem_stream
    return mem_buf


def search_list_value(list, search_value):
    """search value

    Args:
        list (list): list
        search_value (str): search value

    Returns:
        int: listの位置(存在しないときは-1を返す) / Position of list (returns -1 if it does not exist)
    """
    try:
        return list.index(search_value) + 1
    except ValueError:
        return -1
